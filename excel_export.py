"""
Excel export — AllPets Clinic & Beyond Weekly Dashboard
Two-sheet workbook: Dashboard (print-ready) + Raw Data (pivot-table-ready).
"""
import io
from typing import Optional, List, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dashboard_queries import WeeklyDashboard

# ── Design tokens ─────────────────────────────────────────────────────────────
TEAL      = "1B6B72"   # Brand primary
TEAL_DARK = "124850"   # Date band (deeper)
TEAL_BG   = "E8F4F5"   # Column header wash
KPI_BG    = "F0F8F9"   # KPI box fill
ALT_ROW   = "F5FAFA"   # Alternating data row
TOTAL_BG  = "D4E9EC"   # Total row
WHITE     = "FFFFFF"
TEXT      = "1A2526"   # Near-black
MUTED     = "5A7173"   # Secondary text
GREEN     = "1A7A35"   # Growth positive
RED       = "C0392B"   # Growth negative

FONT = "Calibri"

# Growth format: store as decimal (0.082 → 8.2%), arrows auto-applied
# Excel renders ▲/▼ and font is coloured manually in Python
GROWTH_FMT = '"▲ "0.0%;"▼ "0.0%;"-"'
INR_FMT    = '#,##0'
PCT_FMT    = '0.0%'
INT_FMT    = '#,##0'


# ── Style builders ────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _f(size=9, bold=False, color=TEXT, italic=False) -> Font:
    return Font(size=size, bold=bold, color=color, name=FONT, italic=italic)

def _box_border() -> Border:
    s = Side(style="thin", color="B8CECD")
    return Border(left=s, right=s, top=s, bottom=s)

def _row_sep() -> Border:
    return Border(bottom=Side(style="thin", color="D5E6E7"))

def _al(indent=0) -> Alignment:
    return Alignment(horizontal="left", vertical="center", indent=indent)

def _ar() -> Alignment:
    return Alignment(horizontal="right", vertical="center")

def _ac() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


# ── Number helpers ────────────────────────────────────────────────────────────

def _fmt_inr(v: float) -> str:
    """Indian grouping: ₹3,84,715"""
    v = int(round(abs(v)))
    s = str(v)
    if len(s) <= 3:
        return "₹" + s
    rest, last3 = s[:-3], s[-3:]
    groups = []
    while len(rest) > 2:
        groups.append(rest[-2:])
        rest = rest[:-2]
    if rest:
        groups.append(rest)
    groups.reverse()
    return "₹" + ",".join(groups) + "," + last3


def _fmt_short(v: float) -> str:
    """₹5.73Cr or ₹39.7L"""
    if abs(v) >= 1_00_00_000:
        return "₹%.2fCr" % (v / 1_00_00_000)
    if abs(v) >= 1_00_000:
        return "₹%.1fL" % (v / 1_00_000)
    return "₹%s" % "{:,.0f}".format(v)


def _growth_val(pct: Optional[float]) -> Any:
    """Decimal for Excel %, or '—' string for None."""
    return (pct / 100.0) if pct is not None else "—"


# ── Table writer ──────────────────────────────────────────────────────────────

def _table(
    ws, row: int, col: int,
    title: str,
    headers: List[str],
    data: List[List[Any]],
    fmts: List[Optional[str]],
    total: Optional[List[Any]] = None,
    total_fmts: Optional[List[Optional[str]]] = None,
    ncols: Optional[int] = None,
) -> int:
    """Write a titled table; return next available row (with 1-row gap)."""
    nc = ncols or len(headers)
    ec = col + nc - 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=ec)
    tc = ws.cell(row=row, column=col, value=title)
    tc.fill      = _fill(TEAL)
    tc.font      = _f(size=8, bold=True, color=WHITE)
    tc.alignment = _al(indent=1)
    tc.border    = _box_border()
    ws.row_dimensions[row].height = 15
    row += 1

    # ── Column headers ─────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 13
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col + i, value=h)
        c.fill      = _fill(TEAL_BG)
        c.font      = _f(size=8, bold=True, color=TEAL)
        c.alignment = _ar() if i > 0 else _al(indent=1)
        c.border    = _box_border()
    row += 1

    # ── Data rows ──────────────────────────────────────────────────────────
    for di, dr in enumerate(data):
        ws.row_dimensions[row].height = 13
        bg = ALT_ROW if di % 2 == 1 else WHITE
        for i, val in enumerate(dr):
            c = ws.cell(row=row, column=col + i, value=val)
            c.fill      = _fill(bg)
            c.alignment = _ar() if i > 0 else _al(indent=1)
            c.border    = _row_sep()
            fmt = fmts[i] if fmts and i < len(fmts) else None
            # Growth column: arrows via format + font colour via Python
            if fmt == GROWTH_FMT and isinstance(val, float):
                c.number_format = GROWTH_FMT
                if val > 0:
                    c.font = _f(size=9, bold=True, color=GREEN)
                elif val < 0:
                    c.font = _f(size=9, bold=True, color=RED)
                else:
                    c.font = _f(size=9, color=MUTED)
            else:
                c.font = _f(size=9, color=TEXT)
                if fmt and isinstance(val, (int, float)):
                    c.number_format = fmt
        row += 1

    # ── Total row ──────────────────────────────────────────────────────────
    if total is not None:
        ws.row_dimensions[row].height = 13
        tf = total_fmts or fmts
        for i, val in enumerate(total):
            c = ws.cell(row=row, column=col + i, value=val)
            c.fill      = _fill(TOTAL_BG)
            c.font      = _f(size=9, bold=True, color=TEXT)
            c.alignment = _ar() if i > 0 else _al(indent=1)
            c.border    = _box_border()
            fmt = tf[i] if tf and i < len(tf) else None
            if fmt and isinstance(val, (int, float)):
                c.number_format = fmt
        row += 1

    return row + 1   # one blank gap


# ── Raw Data sheet ────────────────────────────────────────────────────────────

def _raw_data_sheet(ws, d: WeeklyDashboard):
    headers = ["Section", "Label", "Sub-Label",
               "This Week", "Prior 3-Wk Avg", "Growth (decimal)", "Share (decimal)"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = _fill(TEAL)
        c.font = Font(bold=True, color=WHITE, name=FONT, size=9)

    rows = []
    ts = d.total_sales
    rows.append(["Revenue", "Total Sales", "", ts.value, ts.prior_value,
                 (ts.variance_pct or 0) / 100, 1.0])
    for s in d.species_split:
        rows.append(["Species", s.species, "", s.this_week_revenue, s.prior_3wk_avg,
                     (s.growth_pct or 0) / 100, s.pct_contribution / 100])
    for cat in d.category_top5:
        rows.append(["Category", cat.category, "", cat.this_week_revenue, cat.prior_3wk_avg,
                     (cat.growth_pct or 0) / 100, cat.pct_contribution / 100])
    for b in d.invoice_count_by_species:
        rows.append(["Bills", b.species, "", b.bill_count, b.prior_3wk_avg,
                     (b.growth_pct or 0) / 100, b.pct_contribution / 100])
    for dn in d.day_night_split:
        rows.append(["Day/Night", dn.time_band, "", dn.revenue, dn.prior_3wk_avg,
                     (dn.growth_pct or 0) / 100, dn.pct_contribution / 100])
    for ls in d.life_stage:
        rows.append(["Life Stage", ls.species, ls.life_stage, ls.revenue, "", 0, 0])
    for iv in d.inventory_by_type:
        rows.append(["Inventory", iv.inventory_type, "", iv.stock_value, "", 0, iv.pct_of_total / 100])
    nv = d.new_vs_existing_revenue
    rows.append(["Customers", "New", "", nv.new_revenue, "", 0, nv.new_pct / 100])
    rows.append(["Customers", "Existing", "", nv.existing_revenue, "", 0, nv.existing_pct / 100])

    for ri, row in enumerate(rows, 2):
        bg = WHITE if ri % 2 == 0 else ALT_ROW
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = _fill(bg)
            c.font = _f(size=9)
            if ci == 6 and isinstance(val, float):
                c.number_format = GROWTH_FMT
                c.font = _f(size=9, bold=True, color=GREEN if val > 0 else (RED if val < 0 else MUTED))
            if ci == 7 and isinstance(val, float):
                c.number_format = PCT_FMT
            if ci in (4, 5) and isinstance(val, (int, float)):
                c.number_format = INR_FMT

    for i, w in enumerate([14, 20, 16, 14, 14, 14, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Main ──────────────────────────────────────────────────────────────────────

def generate_excel(d: WeeklyDashboard) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    wd = wb.create_sheet("Raw Data")
    _raw_data_sheet(wd, d)

    # Page setup — A4 landscape, fit to 1 page wide
    ws.page_setup.orientation  = "landscape"
    ws.page_setup.paperSize    = 9
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.sheet_view.showGridLines = False
    ws.freeze_panes            = "A10"

    # Column widths: left block A-D, spacer E, right block F-J
    for col, w in zip("ABCDEFGHIJ",
                      [22, 13, 9, 13,   # left tables (4 cols)
                       2,               # spacer
                       20, 12, 9, 9, 9]):  # right tables (5 cols)
        ws.column_dimensions[col].width = w

    # ── HEADER BAND ───────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "AllPets Clinic & Beyond"
    c.fill      = _fill(TEAL)
    c.font      = Font(size=20, bold=True, color=WHITE, name=FONT)
    c.alignment = _ac()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value     = "Weekly Performance Dashboard"
    c.fill      = _fill(TEAL)
    c.font      = Font(size=10, color="A8D8DD", name=FONT)
    c.alignment = _ac()
    ws.row_dimensions[2].height = 16

    ws.merge_cells("A3:J3")
    c = ws["A3"]
    c.value     = "📅  " + d.week
    c.fill      = _fill(TEAL_DARK)
    c.font      = Font(size=9, bold=True, color=WHITE, name=FONT)
    c.alignment = _ac()
    ws.row_dimensions[3].height = 14

    ws.row_dimensions[4].height = 5   # gap

    # ── KPI BOXES (rows 5-8) ──────────────────────────────────────────────
    ts  = d.total_sales
    rc  = d.repeat_customer_pct
    inv = sum(r.stock_value for r in d.inventory_by_type)
    inv_pharm    = next((r.stock_value for r in d.inventory_by_type if r.inventory_type == "Pharmacy"), 0)
    inv_nonfood  = next((r.stock_value for r in d.inventory_by_type if r.inventory_type == "Non-Food"), 0)

    kpis = [
        dict(
            icon_label = "🛒  TOTAL SALES",
            value      = ts.value,
            value_fmt  = INR_FMT,
            sub1       = "3-wk avg  " + _fmt_inr(ts.prior_value),
            sub2       = ("%+.1f%% vs prior average" % ts.variance_pct) if ts.variance_pct is not None else "—",
            sub2_color = GREEN if (ts.variance_pct or 0) >= 0 else RED,
            cols       = (1, 3),
        ),
        dict(
            icon_label = "👥  REPEAT CUSTOMERS",
            value      = "%.1f%%" % rc.repeat_pct,
            value_fmt  = "@",
            sub1       = "%d repeat of %d clients this week" % (rc.repeat_customers, rc.total_customers),
            sub2       = "New first-timers: %d" % rc.new_customers,
            sub2_color = MUTED,
            cols       = (4, 7),
        ),
        dict(
            icon_label = "📦  INVENTORY VALUE",
            value      = inv,
            value_fmt  = INR_FMT,
            sub1       = _fmt_short(inv_pharm) + " Pharmacy  ·  " + _fmt_short(inv_nonfood) + " Non-Food",
            sub2       = "Excludes Cytopoint & Lasix phantoms",
            sub2_color = MUTED,
            cols       = (8, 10),
        ),
    ]

    for r_idx, h in {5: 12, 6: 26, 7: 13, 8: 13}.items():
        ws.row_dimensions[r_idx].height = h

    for kpi in kpis:
        sc, ec = kpi["cols"]
        # Background fill + border for all 4 KPI rows
        for r in range(5, 9):
            for c_idx in range(sc, ec + 1):
                cell = ws.cell(row=r, column=c_idx)
                cell.fill   = _fill(KPI_BG)
                cell.border = _box_border()

        ws.merge_cells(start_row=5, start_column=sc, end_row=5, end_column=ec)
        c = ws.cell(row=5, column=sc, value=kpi["icon_label"])
        c.fill = _fill(KPI_BG); c.font = _f(8, bold=True, color=TEAL)
        c.alignment = _al(indent=1); c.border = _box_border()

        ws.merge_cells(start_row=6, start_column=sc, end_row=6, end_column=ec)
        c = ws.cell(row=6, column=sc, value=kpi["value"])
        c.fill = _fill(KPI_BG); c.font = Font(size=18, bold=True, color=TEXT, name=FONT)
        c.number_format = kpi["value_fmt"]
        c.alignment = _al(indent=1); c.border = _box_border()

        ws.merge_cells(start_row=7, start_column=sc, end_row=7, end_column=ec)
        c = ws.cell(row=7, column=sc, value=kpi["sub1"])
        c.fill = _fill(KPI_BG); c.font = _f(8, color=MUTED)
        c.alignment = _al(indent=1); c.border = _box_border()

        ws.merge_cells(start_row=8, start_column=sc, end_row=8, end_column=ec)
        c = ws.cell(row=8, column=sc, value=kpi["sub2"])
        c.fill = _fill(KPI_BG); c.font = _f(8, bold=(kpi["sub2_color"] != MUTED), color=kpi["sub2_color"])
        c.alignment = _al(indent=1); c.border = _box_border()

    ws.row_dimensions[9].height = 6   # gap before tables

    # ── DATA TABLES — left (cols 1-4) and right (cols 6-10) ──────────────
    row_l = 10
    row_r = 10

    # LEFT 1: Species Sales
    row_l = _table(ws, row_l, 1,
        "SPECIES-WISE SALES SPLIT",
        ["Species", "Sales (₹)", "% Share", "vs 3-Wk Avg"],
        [[r.species, r.this_week_revenue, r.pct_contribution / 100, _growth_val(r.growth_pct)]
         for r in d.species_split if r.this_week_revenue > 0],
        [None, INR_FMT, PCT_FMT, GROWTH_FMT],
        total=["Total", sum(r.this_week_revenue for r in d.species_split), 1.0, ""],
        total_fmts=[None, INR_FMT, PCT_FMT, None],
        ncols=4,
    )

    # RIGHT 1: New vs Old Customers
    row_r = _table(ws, row_r, 6,
        "NEW vs OLD CUSTOMERS — BY SPECIES",
        ["Type", "Species", "Clients", "Revenue (₹)"],
        [[r.customer_type, r.species, r.client_count, r.revenue]
         for r in d.new_vs_existing_customers],
        [None, None, INT_FMT, INR_FMT],
        total=["Total", "", d.repeat_customer_pct.total_customers, ""],
        total_fmts=[None, None, INT_FMT, None],
        ncols=5,
    )

    # LEFT 2: Category Split
    row_l = _table(ws, row_l, 1,
        "CATEGORY-WISE SALES — TOP 5",
        ["Category", "Sales (₹)", "% Share", "vs 3-Wk Avg"],
        [[r.category, r.this_week_revenue, r.pct_contribution / 100, _growth_val(r.growth_pct)]
         for r in d.category_top5],
        [None, INR_FMT, PCT_FMT, GROWTH_FMT],
        total=["Total (Top 5)", sum(r.this_week_revenue for r in d.category_top5), "—", ""],
        total_fmts=[None, INR_FMT, None, None],
        ncols=4,
    )

    # RIGHT 2: New vs Old Revenue
    nv = d.new_vs_existing_revenue
    row_r = _table(ws, row_r, 6,
        "TOTAL SALES: NEW vs OLD CUSTOMERS",
        ["Segment", "Revenue (₹)", "Share"],
        [["New Customers", nv.new_revenue, nv.new_pct / 100],
         ["Existing Customers", nv.existing_revenue, nv.existing_pct / 100]],
        [None, INR_FMT, PCT_FMT],
        total=["Total", nv.total_revenue, 1.0],
        total_fmts=[None, INR_FMT, PCT_FMT],
        ncols=5,
    )

    # LEFT 3: Bills
    row_l = _table(ws, row_l, 1,
        "TOTAL BILLS — BY SPECIES",
        ["Species", "Bills (#)", "% Share", "vs 3-Wk Avg"],
        [[r.species, r.bill_count, r.pct_contribution / 100, _growth_val(r.growth_pct)]
         for r in d.invoice_count_by_species if r.bill_count > 0],
        [None, INT_FMT, PCT_FMT, GROWTH_FMT],
        total=["Total", sum(r.bill_count for r in d.invoice_count_by_species), 1.0, ""],
        total_fmts=[None, INT_FMT, PCT_FMT, None],
        ncols=4,
    )

    # RIGHT 3: Life Stage
    row_r = _table(ws, row_r, 6,
        "LIFE STAGE REVENUE — CANINE & FELINE",
        ["Species", "Stage", "Revenue (₹)", "Invoices"],
        [[r.species, r.life_stage, r.revenue, r.invoice_count]
         for r in d.life_stage],
        [None, None, INR_FMT, INT_FMT],
        total=["Total", "", sum(r.revenue for r in d.life_stage),
               sum(r.invoice_count for r in d.life_stage)],
        total_fmts=[None, None, INR_FMT, INT_FMT],
        ncols=5,
    )

    # LEFT 4: Day / Night
    row_l = _table(ws, row_l, 1,
        "DAY vs NIGHT SALES SPLIT",
        ["Time Band", "Sales (₹)", "% Share", "vs 3-Wk Avg"],
        [[r.time_band, r.revenue, r.pct_contribution / 100, _growth_val(r.growth_pct)]
         for r in d.day_night_split],
        [None, INR_FMT, PCT_FMT, GROWTH_FMT],
        total=["Total", sum(r.revenue for r in d.day_night_split), 1.0, ""],
        total_fmts=[None, INR_FMT, PCT_FMT, None],
        ncols=4,
    )

    # RIGHT 4: Inventory
    row_r = _table(ws, row_r, 6,
        "INVENTORY VALUE — BY TYPE",
        ["Category", "Value (₹)", "% Share"],
        [[r.inventory_type, r.stock_value, r.pct_of_total / 100]
         for r in d.inventory_by_type],
        [None, INR_FMT, PCT_FMT],
        total=["Total", sum(r.stock_value for r in d.inventory_by_type), 1.0],
        total_fmts=[None, INR_FMT, PCT_FMT],
        ncols=5,
    )

    # ── Footer ────────────────────────────────────────────────────────────
    last_row = max(row_l, row_r) + 1
    ws.merge_cells(start_row=last_row, start_column=1,
                   end_row=last_row, end_column=10)
    fc = ws.cell(row=last_row, column=1,
                 value="▲ Green = growth vs prior 3-week average  ·  "
                       "▼ Red = decline  ·  Source: DB  ·  "
                       "Inventory excludes Cytopoint & Lasix phantom entries")
    fc.font      = _f(7, color=MUTED, italic=True)
    fc.alignment = _ac()
    ws.row_dimensions[last_row].height = 12

    ws.print_area       = "A1:J%d" % last_row
    ws.print_title_rows = "1:3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
